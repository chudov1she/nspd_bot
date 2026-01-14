"""
Обработчик Excel файлов - формирование выходных файлов с данными.
"""
from pathlib import Path
from typing import List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from loguru import logger

from bot.models.cadastral import RealEstateObject
from bot.config.settings import settings


class ExcelHandlerError(Exception):
    """Исключение при работе с Excel файлами."""
    pass


def create_output_excel(
    results: List[RealEstateObject],
    source_file_path: Optional[Path] = None
) -> Path:
    """
    Создает выходной Excel файл с данными из API.
    
    Args:
        results: Список объектов RealEstateObject с данными
        source_file_path: Путь к исходному файлу (если был загружен)
        
    Returns:
        Путь к созданному файлу
        
    Raises:
        ExcelHandlerError: Если ошибка при создании файла
    """
    try:
        if source_file_path and source_file_path.exists():
            # Если есть исходный файл - используем его как основу
            wb = load_workbook(source_file_path)
            ws = wb.active
            
            # Ищем колонку с кадастровыми номерами
            cadastral_col = _find_cadastral_column(ws)
            
            if cadastral_col:
                # Добавляем новые колонки после существующих
                _add_data_columns(ws, cadastral_col, results)
            else:
                # Если не нашли колонку - создаем новый лист
                ws = wb.create_sheet("Данные из Росреестра")
                _create_full_table(ws, results)
        else:
            # Создаем новый файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Данные из Росреестра"
            _create_full_table(ws, results)
        
        # Сохраняем основной файл (БЕЗ таблицы с картами - она будет в отдельном файле)
        output_dir = settings.OUTPUT_DIR
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Генерируем имя файла (СНГ формат: ДД.ММ.ГГГГ_ЧЧ-ММ-СС, московское время)
        from bot.utils.datetime import strftime_moscow
        timestamp = strftime_moscow("%d.%m.%Y_%H-%M-%S")
        if source_file_path:
            file_name = f"{source_file_path.stem}_результат_{timestamp}.xlsx"
        else:
            file_name = f"росреестр_данные_{timestamp}.xlsx"
        
        output_file = output_dir / file_name
        wb.save(output_file)
        
        logger.info(f"Создан основной файл: {output_file}")
        return output_file
        
    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {e}", exc_info=True)
        raise ExcelHandlerError(f"Ошибка при создании файла: {str(e)}") from e


def _find_cadastral_column(ws) -> Optional[int]:
    """Находит колонку с кадастровыми номерами в существующем файле."""
    # Проверяем первые 10 строк
    for row in range(1, min(11, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                if (cell_str == "кадастровый номер" or 
                    ('кадастровый' in cell_str and 'номер' in cell_str)):
                    return col
    return None


def _add_data_columns(ws, cadastral_col: int, results: List[RealEstateObject]):
    """Добавляет колонки с данными в существующий файл."""
    # Создаем словарь для быстрого поиска данных по кадастровому номеру
    # Используем нормализованные номера для сопоставления
    data_dict = {obj.get_cadastral_number_for_matching(): obj for obj in results}
    
    # Определяем заголовки новых колонок (только необходимые поля)
    headers = [
        "Наименование",
        "Адрес (местоположение)",
        "Общая площадь, кв. м",
        "Категория",
        "Разрешенное использование",
        "Кадастровая стоимость",
        "Передаваемые права",
        "Правообладатель (собственник)",
        "Обременения (ограничения)",
        "Этаж",
        "Статус объекта",
        "Назначение",
        "Дата регистрации права",
        "Дата обновления информации",
        "Старый кадастровый номер",
        "Дата кадастровой стоимости",
        "Карта"
    ]
    
    # Находим первую свободную колонку
    start_col = ws.max_column + 1
    
    # Записываем заголовки
    header_row = 1
    for idx, header in enumerate(headers):
        col = start_col + idx
        cell = ws.cell(header_row, col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Заполняем данные
    for row in range(2, ws.max_row + 1):
        cadastral_value = ws.cell(row, cadastral_col).value
        if cadastral_value:
            # Нормализуем номер для поиска
            cadastral_str = str(cadastral_value).strip().replace(" ", "")
            obj = data_dict.get(cadastral_str)
            
            # Если не нашли точное совпадение, пробуем найти частичное
            if not obj:
                for key, value in data_dict.items():
                    if key.replace(" ", "") == cadastral_str or cadastral_str in key:
                        obj = value
                        break
            
            if obj:
                data = obj.to_dict()
                for idx, header in enumerate(headers):
                    col = start_col + idx
                    
                    # Обрабатываем колонку с картой отдельно
                    if header == "Карта":
                        # Добавляем карту только для земельных участков
                        if obj.is_land_plot() and obj.map_image_path and Path(obj.map_image_path).exists():
                            try:
                                # Добавляем изображение карты
                                img = OpenpyxlImage(obj.map_image_path)
                                img.width = 200
                                img.height = 120
                                
                                # Увеличиваем высоту строки для размещения карты
                                row_height = 110  # единицы Excel
                                ws.row_dimensions[row].height = row_height
                                
                                col_letter = get_column_letter(col)
                                col_width = ws.column_dimensions[col_letter].width or 30
                                
                                # Вычисляем смещение для центрирования по горизонтали
                                # В Excel 1 единица ширины колонки ≈ 7 пикселей
                                col_width_px = col_width * 7
                                offset_x_px = max(0, (col_width_px - img.width) / 2)
                                
                                # Вычисляем смещение для центрирования по вертикали
                                # В Excel 1 единица высоты строки ≈ 1.33 пикселя
                                row_height_px = row_height * 1.33
                                offset_y_px = max(0, (row_height_px - img.height) / 2)
                                
                                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                                from openpyxl.drawing.xdr import XDRPositiveSize2D
                                
                                width_emu = int(img.width * 9525)
                                height_emu = int(img.height * 9525)
                                
                                # Конвертируем смещения в EMU (1 пиксель = 9525 EMU)
                                col_off_emu = int(offset_x_px * 9525)
                                row_off_emu = int(offset_y_px * 9525)
                                
                                img.anchor = OneCellAnchor(
                                    _from=AnchorMarker(
                                        col=col - 1,  # 0-indexed
                                        colOff=col_off_emu,  # Центрирование по горизонтали
                                        row=row - 1,  # 0-indexed
                                        rowOff=row_off_emu  # Центрирование по вертикали
                                    ),
                                    ext=XDRPositiveSize2D(cx=width_emu, cy=height_emu)
                                )
                                
                                ws.add_image(img)
                            except Exception as e:
                                logger.warning(f"Не удалось добавить карту для {obj.cadastral_number}: {e}")
                                # Если ошибка при добавлении карты - пишем текст
                                cell = ws.cell(row, col, value="Ошибка загрузки карты")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            # Для земельных участков без карты - пишем текст
                            if obj.is_land_plot():
                                cell = ws.cell(row, col, value="Карта не найдена")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        # Для не-земельных участков - оставляем пустым
                        continue
                    
                    value = data.get(header, "")
                    # Если есть ошибка, показываем её в поле "Обременения"
                    if obj.has_error() and header == "Обременения (ограничения)":
                        value = f"ОШИБКА: {obj.error}"
                    
                    cell = ws.cell(row, col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Выделяем строки с ошибками
                    if obj.has_error():
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Настраиваем ширину колонок
    column_widths = {
        'Наименование': 20,
        'Адрес (местоположение)': 60,
        'Общая площадь, кв. м': 18,
        'Категория': 25,
        'Разрешенное использование': 25,
        'Кадастровая стоимость': 22,
        'Передаваемые права': 30,
        'Правообладатель (собственник)': 30,
        'Обременения (ограничения)': 30,
        'Карта': 30
    }
    
    for idx, header in enumerate(headers):
        col_letter = get_column_letter(start_col + idx)
        width = column_widths.get(header, 15)
        ws.column_dimensions[col_letter].width = width


def _create_full_table(ws, results: List[RealEstateObject]):
    """Создает полную таблицу с данными (только необходимые поля)."""
    # Заголовки согласно требованиям к отчёту
    headers = [
        "№",
        "Наименование",
        "Кадастровый номер",
        "Адрес (местоположение)",
        "Общая площадь, кв. м",
        "Категория",
        "Разрешенное использование",
        "Кадастровая стоимость",
        "Передаваемые права",
        "Правообладатель (собственник)",
        "Обременения (ограничения)",
        "Этаж",
        "Статус объекта",
        "Назначение",
        "Дата регистрации права",
        "Дата обновления информации",
        "Старый кадастровый номер",
        "Дата кадастровой стоимости",
        "Карта"
    ]
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Заполняем данные
    for row_idx, obj in enumerate(results, start=2):
        data = obj.to_dict()
        # Заполняем номер по порядку
        data["№"] = row_idx - 1
        
        for col_idx, header in enumerate(headers, start=1):
            # Обрабатываем колонку с картой отдельно
            if header == "Карта":
                # Добавляем карту только для земельных участков
                if obj.is_land_plot() and obj.map_image_path and Path(obj.map_image_path).exists():
                    try:
                        # Добавляем изображение карты
                        img = OpenpyxlImage(obj.map_image_path)
                        img.width = 200
                        img.height = 120
                        
                        # Увеличиваем высоту строки для размещения карты
                        row_height = 110  # единицы Excel
                        ws.row_dimensions[row_idx].height = row_height
                        
                        col_letter = get_column_letter(col_idx)
                        col_width = ws.column_dimensions[col_letter].width or 30
                        
                        # Вычисляем смещение для центрирования по горизонтали
                        # В Excel 1 единица ширины колонки ≈ 7 пикселей
                        col_width_px = col_width * 7
                        offset_x_px = max(0, (col_width_px - img.width) / 2)
                        
                        # Вычисляем смещение для центрирования по вертикали
                        # В Excel 1 единица высоты строки ≈ 1.33 пикселя
                        row_height_px = row_height * 1.33
                        offset_y_px = max(0, (row_height_px - img.height) / 2)
                        
                        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                        from openpyxl.drawing.xdr import XDRPositiveSize2D
                        
                        width_emu = int(img.width * 9525)
                        height_emu = int(img.height * 9525)
                        
                        # Конвертируем смещения в EMU (1 пиксель = 9525 EMU)
                        col_off_emu = int(offset_x_px * 9525)
                        row_off_emu = int(offset_y_px * 9525)
                        
                        img.anchor = OneCellAnchor(
                            _from=AnchorMarker(
                                col=col_idx - 1,  # 0-indexed
                                colOff=col_off_emu,  # Центрирование по горизонтали
                                row=row_idx - 1,  # 0-indexed
                                rowOff=row_off_emu  # Центрирование по вертикали
                            ),
                            ext=XDRPositiveSize2D(cx=width_emu, cy=height_emu)
                        )
                        
                        ws.add_image(img)
                    except Exception as e:
                        logger.warning(f"Не удалось добавить карту для {obj.cadastral_number}: {e}")
                        # Если ошибка при добавлении карты - пишем текст
                        cell = ws.cell(row_idx, col_idx, value="Ошибка загрузки карты")
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Для земельных участков без карты - пишем текст
                    if obj.is_land_plot():
                        cell = ws.cell(row_idx, col_idx, value="Карта не найдена")
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Для не-земельных участков - оставляем пустым
                continue
            
            value = data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Выделяем строки с ошибками (если есть ошибка, показываем в поле "Обременения")
            if obj.has_error():
                # Если есть ошибка, показываем её в последнем столбце
                if header == "Обременения (ограничения)":
                    cell.value = f"ОШИБКА: {obj.error}"
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Настраиваем ширину колонок
    column_widths = {
        'A': 8,   # №
        'B': 20,  # Наименование
        'C': 25,  # Кадастровый номер
        'D': 60,  # Адрес (местоположение)
        'E': 18,  # Общая площадь, кв. м
        'F': 25,  # Категория
        'G': 25,  # Разрешенное использование
        'H': 22,  # Кадастровая стоимость
        'I': 30,  # Передаваемые права
        'J': 30,  # Правообладатель (собственник)
        'K': 30,  # Обременения (ограничения)
        'L': 15,  # Этаж
        'M': 18,  # Статус объекта
        'N': 25,  # Назначение
        'O': 22,  # Дата регистрации права
        'P': 25,  # Дата обновления информации
        'Q': 25,  # Старый кадастровый номер
        'R': 25,  # Дата кадастровой стоимости
        'S': 30   # Карта
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


def create_maps_excel(
    land_plots: List[RealEstateObject],
    source_file_path: Optional[Path] = None
) -> Optional[Path]:
    """
    Создает ОТДЕЛЬНЫЙ Excel файл с таблицей карт земельных участков.
    
    Args:
        land_plots: Список земельных участков с картами
        source_file_path: Путь к исходному файлу (для генерации имени)
        
    Returns:
        Путь к созданному файлу или None, если нет участков с картами
        
    Raises:
        ExcelHandlerError: Если ошибка при создании файла
    """
    if not land_plots:
        return None
    
    try:
        # Создаем новый файл только для карт
        wb = Workbook()
        ws = wb.active
        ws.title = "Карты участков"
        
        # Создаем таблицу с картами
        _create_maps_table(ws, land_plots)
        
        # Сохраняем файл
        output_dir = settings.OUTPUT_DIR
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Генерируем имя файла (СНГ формат: ДД.ММ.ГГГГ_ЧЧ-ММ-СС, московское время)
        from bot.utils.datetime import strftime_moscow
        timestamp = strftime_moscow("%d.%m.%Y_%H-%M-%S")
        if source_file_path:
            file_name = f"{source_file_path.stem}_карты_{timestamp}.xlsx"
        else:
            file_name = f"росреестр_карты_{timestamp}.xlsx"
        
        output_file = output_dir / file_name
        wb.save(output_file)
        
        logger.info(f"Создан файл с картами: {output_file}")
        return output_file
        
    except Exception as e:
        logger.error(f"Ошибка при создании файла с картами: {e}", exc_info=True)
        raise ExcelHandlerError(f"Ошибка при создании файла с картами: {str(e)}") from e


def _create_maps_table(ws, land_plots: List[RealEstateObject]):
    """
    Создает таблицу с картами земельных участков на переданном листе.
    
    Структура таблицы:
    - №
    - Наименование
    - Кадастровый номер
    - Инженерные коммуникации
    - Форма (карта участка)
    """
    # Заголовки
    headers = [
        "№",
        "Наименование",
        "Кадастровый номер",
        "Инженерные коммуникации",
        "Форма"
    ]
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Заполняем данные
    for row_idx, obj in enumerate(land_plots, start=2):
        # №
        cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Наименование - всегда "Земельный участок"
        cell = ws.cell(row=row_idx, column=2, value="Земельный участок")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Кадастровый номер
        cell = ws.cell(row=row_idx, column=3, value=obj.cadastral_number)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Инженерные коммуникации
        engineering = obj.engineering_communications or "Отсутствуют"
        cell = ws.cell(row=row_idx, column=4, value=engineering)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Форма - добавляем карту и текст формы
        if obj.map_image_path and Path(obj.map_image_path).exists():
            try:
                # Добавляем изображение карты (прямоугольная форма: ширина больше высоты)
                img = OpenpyxlImage(obj.map_image_path)
                # Прямоугольная карта: 200x120 пикселей (ширина больше высоты)
                img.width = 200
                img.height = 120
                
                # Увеличиваем высоту строки для размещения карты и текста под ней
                # Высота строки: 110 единиц Excel = 147 пикселей
                ws.row_dimensions[row_idx].height = 110
                
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(5)  # Колонка E
                
                # Получаем ширину колонки для центрирования изображения
                # В Excel 1 единица ширины колонки ≈ 7 пикселей (зависит от шрифта)
                col_width = ws.column_dimensions[col_letter].width or 30  # Ширина колонки в единицах Excel
                col_width_px = col_width * 7  # Примерная ширина в пикселях
                
                # Вычисляем смещение для центрирования изображения по горизонтали
                # Изображение шириной 200px, колонка col_width_px
                # Смещение = (ширина колонки - ширина изображения) / 2
                offset_x = max(0, (col_width_px - 200) / 2)
                
                # Добавляем изображение с центрированием через anchor
                # Используем ячейку как якорь и добавляем смещение для центрирования
                # В openpyxl изображение позиционируется относительно верхнего левого угла ячейки
                # Для центрирования используем смещение в единицах EMU (1 пиксель = 9525 EMU)
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
                
                # Размеры изображения в EMU (1 пиксель = 9525 EMU)
                width_emu = int(img.width * 9525)
                height_emu = int(img.height * 9525)
                
                # Создаем anchor для позиционирования изображения по центру сверху
                # Изображение будет в верхней части ячейки, центрировано горизонтально
                img.anchor = OneCellAnchor(
                    _from=AnchorMarker(
                        col=4,  # Колонка E (0-indexed: 4)
                        colOff=int(offset_x * 9525),  # Смещение для центрирования
                        row=row_idx - 1,  # Строка (0-indexed)
                        rowOff=0  # Верх ячейки
                    ),
                    ext=XDRPositiveSize2D(cx=width_emu, cy=height_emu)  # Размеры изображения
                )
                
                # Добавляем изображение на лист
                ws.add_image(img)
                
                # Добавляем текстовое описание формы в ту же ячейку (под картой)
                # Текст будет по центру внизу ячейки, под изображением
                form_text = obj.form or "Близка к прямоугольной"
                cell = ws.cell(row=row_idx, column=5)
                cell.value = form_text
                # Выравнивание: по центру горизонтально, по низу вертикально
                # Это разместит текст по центру внизу ячейки, под изображением
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                
            except Exception as e:
                logger.warning(f"Не удалось добавить карту для {obj.cadastral_number}: {e}")
                # Если не удалось добавить карту, просто пишем текст формы
                form_text = obj.form or "Близка к прямоугольной"
                cell = ws.cell(row=row_idx, column=5, value=form_text)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            # Если карты нет, просто пишем текст формы
            form_text = obj.form or "Близка к прямоугольной"
            cell = ws.cell(row=row_idx, column=5, value=form_text)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Настраиваем ширину колонок
    column_widths = {
        'A': 8,   # №
        'B': 25,  # Наименование
        'C': 30,  # Кадастровый номер
        'D': 35,  # Инженерные коммуникации
        'E': 30   # Форма (карта) - увеличена для прямоугольной карты 200x120
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Настраиваем выравнивание для всех ячеек - все по центру
    for row in range(2, ws.max_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            # Все ячейки выровнены по центру (и горизонтально, и вертикально)
            if col == 5:  # Колонка "Форма" - текст по центру снизу (для текста под картой)
                # Для колонки с картой текст должен быть снизу, но по центру горизонтально
                # Это уже установлено выше при добавлении текста, но переустанавливаем для уверенности
                if cell.value and isinstance(cell.value, str):  # Если есть текстовое значение
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                # Все остальные ячейки (№, Наименование, Кадастровый номер, Инженерные коммуникации) - по центру
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    logger.info(f"Создана таблица с картами для {len(land_plots)} участков")

