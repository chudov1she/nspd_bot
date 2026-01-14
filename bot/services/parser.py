"""
Парсер кадастровых номеров из текста и файлов.
"""
import re
from pathlib import Path
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from loguru import logger

from bot.utils.validators import CADASTRAL_PATTERN, normalize_cadastral_number


class CadastralParserError(Exception):
    """Исключение при парсинге кадастровых номеров."""
    pass


def extract_cadastral_numbers_from_text(text: str) -> List[str]:
    """
    Извлекает кадастровые номера из текста с помощью регулярного выражения.
    
    Args:
        text: Текст для парсинга
        
    Returns:
        Список уникальных кадастровых номеров
        
    Examples:
        >>> extract_cadastral_numbers_from_text("78:38:0022629:1115")
        ['78:38:0022629:1115']
        >>> extract_cadastral_numbers_from_text("78:38:0022629:1115, 78:38:0022629:1006")
        ['78:38:0022629:1115', '78:38:0022629:1006']
    """
    if not text:
        return []
    
    # Ищем все совпадения паттерна
    matches = CADASTRAL_PATTERN.findall(text)
    
    # Нормализуем и убираем дубликаты
    normalized_numbers = []
    seen = set()
    
    for match in matches:
        normalized = normalize_cadastral_number(match)
        if normalized and normalized not in seen:
            normalized_numbers.append(normalized)
            seen.add(normalized)
    
    logger.info(f"Извлечено {len(normalized_numbers)} кадастровых номеров из текста")
    
    return normalized_numbers


def extract_cadastral_numbers_from_excel(
    file_path: Path | str,
    sheet_name: Optional[str] = None,
    search_all_cells: bool = True
) -> List[str]:
    """
    Извлекает кадастровые номера из Excel файла.
    
    По умолчанию ищет кадастровые номера во всех ячейках всех листов.
    Если search_all_cells=False, ищет только в колонке "кадастровый номер".
    
    Args:
        file_path: Путь к Excel файлу
        sheet_name: Название листа (если None, обрабатываются все листы)
        search_all_cells: Если True, ищет во всех ячейках, иначе только в колонке "кадастровый номер"
        
    Returns:
        Список уникальных кадастровых номеров
        
    Raises:
        CadastralParserError: Если файл не найден или ошибка чтения
        
    Examples:
        >>> numbers = extract_cadastral_numbers_from_excel("input.xlsx")
        >>> numbers
        ['78:38:0022629:1115', '78:38:0022629:1006']
    """
    file_path = Path(file_path)
    
    # Проверка существования файла
    if not file_path.exists():
        raise CadastralParserError(f"Файл не найден: {file_path}")
    
    # Проверка расширения
    if file_path.suffix.lower() not in ['.xlsx', '.xls']:
        raise CadastralParserError(
            f"Неподдерживаемый формат файла: {file_path.suffix}. "
            "Поддерживаются только .xlsx и .xls"
        )
    
    try:
        # Загружаем книгу
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        all_numbers = []
        seen = set()
        
        # Определяем какие листы обрабатывать
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise CadastralParserError(
                    f"Лист '{sheet_name}' не найден. "
                    f"Доступные листы: {', '.join(wb.sheetnames)}"
                )
            sheets_to_process = [sheet_name]
        else:
            sheets_to_process = wb.sheetnames
        
        logger.info(
            f"Обработка {len(sheets_to_process)} лист(ов) из файла {file_path.name}: "
            f"{', '.join(sheets_to_process)}"
        )
        
        # Обрабатываем каждый лист
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            logger.info(f"Обработка листа '{ws.title}' (строк: {ws.max_row}, колонок: {ws.max_column})")
            
            sheet_numbers_before = len(all_numbers)
            
            if search_all_cells:
                # Ищем во всех ячейках листа
                sheet_numbers = _extract_from_all_cells(ws, seen)
            else:
                # Ищем только в колонке "кадастровый номер" (старый способ)
                sheet_numbers = _extract_from_column(ws, seen)
            
            all_numbers.extend(sheet_numbers)
            sheet_numbers_after = len(all_numbers)
            new_numbers_count = sheet_numbers_after - sheet_numbers_before
            
            logger.info(
                f"Найдено {new_numbers_count} новых номеров на листе '{ws.title}' "
                f"(всего уникальных: {len(all_numbers)})"
            )
        
        wb.close()
        
        logger.info(
            f"Извлечено {len(all_numbers)} уникальных кадастровых номеров "
            f"из файла {file_path.name}"
        )
        
        if not all_numbers:
            raise CadastralParserError(
                "В файле не найдено ни одного валидного кадастрового номера. "
                "Проверьте формат данных в файле."
            )
        
        return all_numbers
        
    except CadastralParserError:
        raise
    except Exception as e:
        logger.error(f"Ошибка при чтении Excel файла {file_path}: {e}")
        raise CadastralParserError(
            f"Ошибка при чтении файла: {str(e)}"
        ) from e


def _extract_from_all_cells(ws, seen: set) -> List[str]:
    """
    Извлекает кадастровые номера из всех ячеек листа.
    
    Args:
        ws: Рабочий лист
        seen: Множество уже найденных номеров (для исключения дубликатов)
        
    Returns:
        Список новых уникальных кадастровых номеров
    """
    numbers = []
    
    # Проходим по всем строкам
    for row in range(1, ws.max_row + 1):
        # Проходим по всем колонкам в строке
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            
            # Пропускаем пустые ячейки
            if cell.value is None:
                continue
            
            # Преобразуем значение в строку
            cell_str = str(cell.value).strip()
            
            # Пропускаем пустые строки
            if not cell_str:
                continue
            
            # Извлекаем кадастровые номера из содержимого ячейки
            extracted = extract_cadastral_numbers_from_text(cell_str)
            
            # Добавляем только новые номера
            for num in extracted:
                if num not in seen:
                    numbers.append(num)
                    seen.add(num)
    
    return numbers


def _extract_from_column(ws, seen: set, column_name: str = "кадастровый номер") -> List[str]:
    """
    Извлекает кадастровые номера из колонки с указанным именем (старый способ).
    
    Args:
        ws: Рабочий лист
        seen: Множество уже найденных номеров
        column_name: Название колонки для поиска
        
    Returns:
        Список новых уникальных кадастровых номеров
    """
    numbers = []
    
    # Ищем заголовок колонки
    header_row = 1
    cadastral_col = None
    column_name_lower = column_name.lower().strip()
    
    # Проверяем первую строку на наличие нужной колонки
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value:
            cell_str = str(cell_value).lower().strip()
            # Проверяем точное совпадение или частичное (содержит ключевые слова)
            if (cell_str == column_name_lower or 
                'кадастровый' in cell_str and 'номер' in cell_str):
                cadastral_col = col
                logger.info(
                    f"Найдена колонка '{cell_value}' в позиции "
                    f"{get_column_letter(col)}"
                )
                break
    
    if not cadastral_col:
        # Пробуем найти в других строках (может быть заголовок не в первой строке)
        for row in range(1, min(10, ws.max_row + 1)):  # Проверяем первые 10 строк
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                if cell_value:
                    cell_str = str(cell_value).lower().strip()
                    if (cell_str == column_name_lower or 
                        'кадастровый' in cell_str and 'номер' in cell_str):
                        cadastral_col = col
                        header_row = row
                        logger.info(
                            f"Найдена колонка '{cell_value}' в строке {row}, "
                            f"позиция {get_column_letter(col)}"
                        )
                        break
            if cadastral_col:
                break
    
    if not cadastral_col:
        raise CadastralParserError(
            f"Колонка '{column_name}' не найдена в файле. "
            "Убедитесь, что файл содержит колонку с названием "
            "'кадастровый номер' или похожим."
        )
    
    # Извлекаем номера из колонки
    start_row = header_row + 1
    
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row, cadastral_col).value
        
        if not cell_value:
            continue
        
        # Пробуем нормализовать значение
        cell_str = str(cell_value).strip()
        
        # Если это уже кадастровый номер - добавляем
        normalized = normalize_cadastral_number(cell_str)
        if normalized and normalized not in seen:
            numbers.append(normalized)
            seen.add(normalized)
            continue
        
        # Если нет, пробуем извлечь из текста (на случай если в ячейке текст с номером)
        extracted = extract_cadastral_numbers_from_text(cell_str)
        for num in extracted:
            if num not in seen:
                numbers.append(num)
                seen.add(num)
    
    return numbers
